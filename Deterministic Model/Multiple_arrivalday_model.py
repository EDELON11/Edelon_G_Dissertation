# Import openyxl module
import openpyxl
import gurobipy as grb
from gurobipy import GRB
import time

start_time = time.time()

# Open workbook from ExcelData

workbook = openpyxl.load_workbook("testData1_chap2.xlsx")

# Define variable to read the active sheet:
worksheet = workbook.active

# Import openyxl module
import openpyxl


#The max Number that can be accepted by the hotel
Max_NumberOfLenght = 14
ArrivalDates = 360
StartDate = 1


demands=[];
for i in range(126*ArrivalDates):
    demands.append(0);
    
demands_to_8=[];
for i in range(560*ArrivalDates):
    demands_to_8.append(0);


#Demands Function - Section 2.2.2.6 (Explained)

for p in range(1, worksheet.max_row):
    for length in worksheet.iter_cols(4, 4):
        for Room_class in worksheet.iter_cols(2, 2):
            for ArrivalDate in worksheet.iter_cols(3,3):
                for adults in worksheet.iter_cols(5, 5):
                    for children in worksheet.iter_cols(6, 6):
                        
                        if ArrivalDate[p].value <= ArrivalDates and ArrivalDate[p].value>= StartDate:
                            
                            if Room_class[p].value <= 3:
                                demand_index = ((Max_NumberOfLenght * ArrivalDates * 3)*(Room_class[p].value-1))\
                                    +(3* ArrivalDates *(length[p].value-1))+adults[p].value + 3*(ArrivalDate[p].value-1)
                                if (adults[p].value == 1 and children[p].value == 0):
                                    demand_index=demand_index-1
                                
                                demands[demand_index] = demands[demand_index]+1
                                
                                
                                
                            elif Room_class[p].value >= 4 and Room_class[p].value <= 8 :
                                demand_index = ((Max_NumberOfLenght * ArrivalDates * 8)*(Room_class[p].value-4)\
                                                +(8*ArrivalDates*(length[p].value-1))+(4*(adults[p].value-1)\
                                                +children[p].value))+ 8*(ArrivalDate[p].value-1)
                                    
                                if (adults[p].value == 3 and children[p].value == 0):
                                    demand_index=demand_index-1
                                
                                demands_to_8[demand_index] = demands_to_8[demand_index]+1

def SpecificDemand(c,a,l,na,nc):
    
    if c <= 3:
        demand_index = ((Max_NumberOfLenght * ArrivalDates * 3)*(c-1))+(3*ArrivalDates*(l-1))+na+(3*(a-1))
        
        if (na == 1 and nc == 0):
            demand_index=demand_index-1
            
        return demands[demand_index]
        
    elif c >= 4 and c <= 8 :
        demand_index = ((Max_NumberOfLenght * ArrivalDates * 8)*(c-4))+(8*ArrivalDates*(l-1))+(4*(na-1)+nc)+(8*(a-1))
        if (na == 3 and nc == 0):
            demand_index=demand_index-1
        
        return demands_to_8[demand_index]
    
#Season indicating its Room class & Base rate

def GetLowRate(Room_class):
    switcher={
        1:90,
        2:105,
        3:125,
        4:140,
        5:150,
        6:165,
        7:200,
        8:230,
        }
    
    return switcher.get(Room_class,"nothing")

def GetShoulderRate(Room_class):
    switcher={
        1:105,
        2:120,
        3:140,
        4:155,
        5:165,
        6:180,
        7:215,
        8:240,
        }
    
    return switcher.get(Room_class,"nothing")

def GetHighRate(Room_class):
    switcher={
        1:120,
        2:135,
        3:155,
        4:170,
        5:180,
        6:195,
        7:225,
        8:250,
        }
    
    return switcher.get(Room_class,"nothing")

def GetVeryHighRate(Room_class):
    switcher={
        1:130,
        2:145,
        3:160,
        4:175,
        5:185,
        6:200,
        7:235,
        8:260,
        }
    
    return switcher.get(Room_class,"nothing")
        
       
def GetRev(c,a,l,na,nc):
    na= int(na);
    nc = int(nc);
    a = int(a);
    c=int(c);
    number_of_persons = na+nc;
    
    #if in july and aug
    if a >= 182 and a<=243:
        BaseRate = GetVeryHighRate(c)
    #if in sept or june
    elif (a >= 244 and a<=273) or (a>=152 and a<=181):
        BaseRate = GetHighRate(c)
    # from nov till 20th dec or from 7th jan till 10th april
    elif(a>= 305 and a<= 354) or (a>=7 and a<=100):
        BaseRate= GetLowRate(c)
    #from 21th dec till 6th jan or october or from 11th april till may
    elif (a>=1 and a<=6) or (a>=355 and a<=365) or (a>=274 and a<=304) or (a>=101 and a<=151):
        BaseRate = GetShoulderRate(c)
        
    if number_of_persons <=2:
        return l*BaseRate;
    elif number_of_persons > 2:
        if na == 3:
            return (l*BaseRate)+((0.75*(0.5*BaseRate))*l);
        elif na == 2:
            return (l*BaseRate)+(l*((0.5*(0.5*BaseRate))*nc));
        elif na == 1:
            return (l*BaseRate)+((0.5*(0.5*BaseRate))*(number_of_persons-2)*l);

#Defining the Model
model = grb.Model(name='Revenue_Managment')

#Define the parameters
C = [i for i in range(1,9)]
A = [i for i in range(1,4)]
J = [i for i in range(1,4)]
L = [i for i in range(1,15)]
N1 = [i for i in range(1,4)]
N2 = [i for i in range(1,9)]
ubs = [5,2,3,2,1,2,1,1]

#Defining Decision variable
x = {}
optimizedVariables=[]
VariablesChosen=[]
Revenue=[]


#GetSum Function - Section 2.2.2.7 (Explained)
def getSum(day_index,c):
    expr = 0
    counter = 1
    if c<=3:
        if day_index <=14:
            for a in range(day_index,0,-1):
                for L in range(1,counter+1):
                    if L+1 > counter:
                        for l in range(L,14+1):
                            for n in N1:
                                expr += x[c,a,l,n]
                counter=counter+1
            
        else:
            for a in range(day_index,day_index-14,-1):
                for L in range(1,counter+1):
                    if L+1 > counter:
                        for l in range(L,14+1):
                            for n in N1:
                                expr += x[c,a,l,n]
                counter=counter+1
                
    elif c>=4 and c<=8:    
        if day_index <=14:
            for a in range(day_index,0,-1):
                for L in range(1,counter+1):
                    if L+1 > counter:
                        for l in range(L,14+1):
                            for n in N2:
                                expr += x[c,a,l,n]
                counter=counter+1
                
        else:
            for a in range(day_index,day_index-14,-1):
                for L in range(1,counter+1):
                    if L+1 > counter:
                        for l in range(L,14+1):
                            for n in N2:
                                expr += x[c,a,l,n]
                counter=counter+1
           
    return expr
    


#The final model that works for the whole year
def FinalModel():
    for c in C:
        if c<=3:
            for a in range(1,ArrivalDates+1):
                for l in L:
                    for n in N1:
                        if n==1:
                            x[c,a,l,n] = model.addVar(vtype=grb.GRB.INTEGER,lb=0,name='x'+str(c)+str(a)+str(l)+'(1,0)')
                        elif n==2:
                            x[c,a,l,n] = model.addVar(vtype=grb.GRB.INTEGER,lb=0,name='x'+str(c)+str(a)+str(l)+'(1,1)')
                        elif n==3:
                            x[c,a,l,n] = model.addVar(vtype=grb.GRB.INTEGER,lb=0,name='x'+str(c)+str(a)+str(l)+'(2,0)')
        elif c>=4 and c <= 8:
            for a in range(1,ArrivalDates+1):
                for l in L:  
                    for n in N2:
                        if n == 1:
                            x[c,a,l,n] = model.addVar(vtype=grb.GRB.INTEGER,lb=0,name='x'+str(c)+str(a)+str(l)+'(1,0)')
                        elif n==2:
                            x[c,a,l,n] = model.addVar(vtype=grb.GRB.INTEGER,lb=0,name='x'+str(c)+str(a)+str(l)+'(1,1)')
                        elif n==3:
                            x[c,a,l,n] = model.addVar(vtype=grb.GRB.INTEGER,lb=0,name='x'+str(c)+str(a)+str(l)+'(1,2)')
                        elif n==4:
                            x[c,a,l,n] = model.addVar(vtype=grb.GRB.INTEGER,lb=0,name='x'+str(c)+str(a)+str(l)+'(1,3)')
                        elif n==5:
                            x[c,a,l,n] = model.addVar(vtype=grb.GRB.INTEGER,lb=0,name='x'+str(c)+str(a)+str(l)+'(2,0)')
                        elif n==6:
                            x[c,a,l,n] = model.addVar(vtype=grb.GRB.INTEGER,lb=0,name='x'+str(c)+str(a)+str(l)+'(2,1)')
                        elif n==7:
                            x[c,a,l,n] = model.addVar(vtype=grb.GRB.INTEGER,lb=0,name='x'+str(c)+str(a)+str(l)+'(2,2)')
                        elif n==8:
                            x[c,a,l,n] = model.addVar(vtype=grb.GRB.INTEGER,lb=0,name='x'+str(c)+str(a)+str(l)+'(3,0)')
    
    for c in C:
        for a in range(1,ArrivalDates+1):
            if c<=3:
                    model.addConstr(getSum(a,c) <= (ubs[c-1]))
            elif c>=4 and c<=8:
                    model.addConstr(getSum(a,c) <= (ubs[c-1]))
                      
    for c in C:
        if c<=3:
            for a in range(1,ArrivalDates+1):
                for l in L:
                    for n in N1:
                        if n==1:
                            model.addConstr(x[c,a,l,n] <= SpecificDemand(c,a,l, 1, 0))
                        elif n==2:
                            model.addConstr(x[c,a,l,n] <= SpecificDemand(c,a,l, 1, 1))
                        elif n==3:
                            model.addConstr(x[c,a,l,n] <= SpecificDemand(c,a,l, 2, 0))
                
        
        elif c>=4 and c <= 8:
            for a in range(1,ArrivalDates+1):
                for l in L:
                    for n in N2:
                        if n==1:
                            model.addConstr(x[c,a,l,n] <= SpecificDemand(c,a,l, 1, 0))
                        elif n==2:
                            model.addConstr(x[c,a,l,n] <= SpecificDemand(c,a,l, 1, 1))
                        elif n==3:
                            model.addConstr(x[c,a,l,n] <= SpecificDemand(c,a,l, 1, 2))
                        elif n==4:
                            model.addConstr(x[c,a,l,n] <= SpecificDemand(c,a,l, 1, 3))
                        elif n==5:
                            model.addConstr(x[c,a,l,n] <= SpecificDemand(c,a,l, 2, 0))
                        elif n==6:
                            model.addConstr(x[c,a,l,n] <= SpecificDemand(c,a,l, 2, 1))
                        elif n==7:
                            model.addConstr(x[c,a,l,n] <= SpecificDemand(c,a,l, 2, 2))
                        elif n==8: 
                            model.addConstr(x[c,a,l,n] <= SpecificDemand(c,a,l, 3, 0))

                
    obj = 0
    for c in C:
        if c<=3:
            for a in range(1,ArrivalDates+1):
                for l in L:
                    for n in N1:
                        if n==1:
                            obj = obj + (x[c,a,l,n] * GetRev(c,a,l, 1, 0))
                        elif n==2:
                            obj = obj + (x[c,a,l,n] * GetRev(c,a,l, 1, 1))
                        elif n==3:
                            obj = obj + (x[c,a,l,n] * GetRev(c,a,l, 2, 0))
            
        elif c>=4 and c <= 8:
            for a in range(1,ArrivalDates+1):
                for l in L:
                    for n in N2:
                        if n==1:
                            obj = obj + (x[c,a,l,n] * GetRev(c,a,l, 1, 0))
                        elif n==2:
                            obj = obj + (x[c,a,l,n] * GetRev(c,a,l, 1, 1))
                        elif n==3:
                            obj = obj + (x[c,a,l,n] * GetRev(c,a,l, 1, 2))
                        elif n==4:
                            obj = obj + (x[c,a,l,n] * GetRev(c,a,l, 1, 3))
                        elif n==5:
                            obj = obj + (x[c,a,l,n] * GetRev(c,a,l, 2, 0))
                        elif n==6:
                            obj = obj + (x[c,a,l,n] * GetRev(c,a,l, 2, 1))
                        elif n==7:
                            obj = obj + (x[c,a,l,n] * GetRev(c,a,l, 2, 2))
                        elif n==8:
                            obj = obj + (x[c,a,l,n] * GetRev(c,a,l, 3, 0))
                                
                            
    
    model.setParam('MIPGap', 0)
    model.setObjective(obj,GRB.MAXIMIZE)
    model.update()
    model.optimize()
    print("Revenue: €",model.objVal)
    print("\n\n")
    
    #Print data extracted
    for v in model.getVars():
        if v.x != 0:
            print(v.varName, v.x)
            
FinalModel()

model.setParam("OutputFlag", 1)
model.setParam("LogToConsole", 1)

end_time = time.time()
duration = end_time - start_time

minutes, seconds = divmod(duration, 60)
print(f"\nThe Gurobi optimization took {int(minutes)} minutes and {seconds:.2f} seconds to complete.")
                   

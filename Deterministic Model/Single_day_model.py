# Import openyxl module
import openpyxl
import gurobipy as grb
from gurobipy import GRB
import time

start_time = time.time()

# Import openyxl module

# Define variable to load the workbook
workbook = openpyxl.load_workbook("testData_chap5.xlsx")

# Define variable to read the active sheet:
worksheet = workbook.active

demands=[];
for i in range(126):
    demands.append(0);
    
demands_to_8=[];
for i in range(560):
    demands_to_8.append(0);
    
Max_NumberOfLenght = 14

#Arrival-Date can be changed from here
a=1
    
#Demands Function - Section 2.2.2.2 (Explained)
for p in range(1, worksheet.max_row):
    for length in worksheet.iter_cols(4, 4):
        for Room_class in worksheet.iter_cols(2, 2):
            for adults in worksheet.iter_cols(5, 5):
                for children in worksheet.iter_cols(6, 6):
                    
                    if Room_class[p].value <= 3:
                        demand_index = ((Max_NumberOfLenght * 3)*(Room_class[p].value-1))\
                            +(3*(length[p].value-1))+adults[p].value
                        if (adults[p].value == 1 and children[p].value == 0):
                            demand_index=demand_index-1
                        
                        demands[demand_index] = demands[demand_index]+1
                        
                    elif Room_class[p].value >= 4 and Room_class[p].value <= 8 :
                        demand_index = ((Max_NumberOfLenght * 8)*(Room_class[p].value-4)\
                                        +(8*(length[p].value-1))+(4*(adults[p].value-1)+children[p].value))
                        if (adults[p].value == 3 and children[p].value == 0):
                            demand_index=demand_index-1
                        
                        demands_to_8[demand_index] = demands_to_8[demand_index]+1
                        
                        
def sum_list(l):
    sum = 0
    for x in l:
        sum += x
    return sum


        
print(sum_list(demands))
print(sum_list(demands_to_8))


def SpecificDemand(c,a,l,na,nc):
    
    if c <= 3:
        demand_index = ((Max_NumberOfLenght * 3)*(c-1))+(3*(l-1))+na
        if (na == 1 and nc == 0):
            demand_index=demand_index-1
            
        return demands[demand_index]
        
    elif c >= 4 and c <= 8 :
        demand_index = ((Max_NumberOfLenght * 8)*(c-4)+(8*(l-1))+(4*(na-1)+nc))
        if (na == 3 and nc == 0):
            demand_index=demand_index-1
        
        return demands_to_8[demand_index]
    

def GetBaseRate(Room_class):
    switcher={
        1:105,
        2:120,
        3:140,
        4:155,
        5:165,
        6:180,
        7:215,
        8:240,
        }
    
    return switcher.get(Room_class,"nothing")
        
#Explained in Section 2.2.2.3       
def GetRev(c,l,na,nc):
    na= int(na);
    nc = int(nc);
    numberpersons = na+nc;
    BaseRate = GetBaseRate(int(c))
    if numberpersons <=2:
        return l*BaseRate;
    elif numberpersons > 2:
        if na == 3:
            return (l*BaseRate)+((0.75*(0.5*BaseRate))*l);
        elif na == 2:
            return (l*BaseRate)+(l*((0.5*(0.5*BaseRate))*nc));
        elif na == 1:
            return (l*BaseRate)+((0.5*(0.5*BaseRate))*(numberpersons-2)*l);

print("Finish")

#Defining the Model
model = grb.Model(name='Revenue_Managment')

#Define the parameters

C = [i for i in range(1,9)]
L = [i for i in range(1,15)]
N1 = [i for i in range(1,4)]
N2 = [i for i in range(1,9)]
ubs = [5,2,3,2,1,2,1,1]


#Defining Decision variable
x = {}
for c in C:
    if c<=3:
        for l in L:
            for n in N1:
                if n == 1:
                    x[c,a,l,n] = model.addVar(vtype=grb.GRB.INTEGER,lb=0,name='x'+str(c)+str(a)+str(l)+'(1,0)')
                elif n == 2:
                    x[c,a,l,n] = model.addVar(vtype=grb.GRB.INTEGER,lb=0,name='x'+str(c)+str(a)+str(l)+'(1,1)')
                elif n == 3:
                    x[c,a,l,n] = model.addVar(vtype=grb.GRB.INTEGER,lb=0,name='x'+str(c)+str(a)+str(l)+'(2,0)')                                                                          
    elif c>=4 and c <= 8:
        for l in L:  
            for n in N2:
                if n == 1:
                    x[c,a,l,n] = model.addVar(vtype=grb.GRB.INTEGER,lb=0,name='x'+str(c)+str(a)+str(l)+'(1,0)')
                elif n==2:
                    x[c,a,l,n] = model.addVar(vtype=grb.GRB.INTEGER,lb=0,name='x'+str(c)+str(a)+str(l)+'(1,1)')
                elif n==3:
                    x[c,a,l,n] = model.addVar(vtype=grb.GRB.INTEGER,lb=0,name='x'+str(c)+str(a)+str(l)+'(1,2)')
                elif n == 4:
                    x[c,a,l,n] = model.addVar(vtype=grb.GRB.INTEGER,lb=0,name='x'+str(c)+str(a)+str(l)+'(1,3)')
                elif n==5:
                    x[c,a,l,n] = model.addVar(vtype=grb.GRB.INTEGER,lb=0,name='x'+str(c)+str(a)+str(l)+'(2,0)')
                elif n==6:
                    x[c,a,l,n] = model.addVar(vtype=grb.GRB.INTEGER,lb=0,name='x'+str(c)+str(a)+str(l)+'(2,1)')
                elif n==7:
                    x[c,a,l,n] = model.addVar(vtype=grb.GRB.INTEGER,lb=0,name='x'+str(c)+str(a)+str(l)+'(2,2)')
                elif n==8:
                    x[c,a,l,n] = model.addVar(vtype=grb.GRB.INTEGER,lb=0,name='x'+str(c)+str(a)+str(l)+'(3,0)')
         
    
#Defining the constraints
                
# Constraint 1
#upperbounds of the unallocated rooms
for c in C:
    if c<=3:
        model.addConstr(grb.quicksum(x[c,a,l,n] for l in L for n in N1) <= ubs[c-1])
    elif c>=4 and c <= 8:
        model.addConstr(grb.quicksum(x[c,a,l,n] for l in L for n in N2) <= ubs[c-1])
        

# Constraint 2
#Rooms allocated smaller than the demands 
for c in C:
    if c<=3:
        for l in L:
            for n in N1:
                if n==1:
                    model.addConstr(x[c,a,l,n] <= SpecificDemand(c,a, l, 1, 0))
                elif n==2:
                    model.addConstr(x[c,a,l,n] <= SpecificDemand(c,a, l, 1, 1))
                elif n==3:
                    model.addConstr(x[c,a,l,n] <= SpecificDemand(c,a, l, 2, 0))
    elif c>=4 and c <= 8:
        for l in L:
            for n in N2:
                if n==1:
                    model.addConstr(x[c,a,l,n] <= SpecificDemand(c,a, l, 1, 0))
                elif n==2:
                    model.addConstr(x[c,a,l,n] <= SpecificDemand(c,a, l, 1, 1))
                elif n==3:
                    model.addConstr(x[c,a,l,n] <= SpecificDemand(c,a, l, 1, 2))
                elif n==4:
                    model.addConstr(x[c,a,l,n] <= SpecificDemand(c,a, l, 1, 3))
                elif n==5:
                    model.addConstr(x[c,a,l,n] <= SpecificDemand(c,a, l, 2, 0))
                elif n==6:
                    model.addConstr(x[c,a,l,n] <= SpecificDemand(c,a, l, 2, 1))
                elif n==7:
                    model.addConstr(x[c,a,l,n] <= SpecificDemand(c,a, l, 2, 2))
                elif n==8:
                    model.addConstr(x[c,a,l,n] <= SpecificDemand(c,a, l, 3, 0))


#Defining the objective function 

obj = 0
for c in C:
    if c<=3:
        for l in L:
            for n in N1:
                if n==1:
                    obj = obj + (x[c,a,l,n] * GetRev(c, l, 1, 0))
                elif n==2:
                    obj = obj + (x[c,a,l,n] * GetRev(c, l, 1, 1))
                elif n==3:
                    obj = obj + (x[c,a,l,n] * GetRev(c, l, 2, 0))
    elif c>=4 and c <= 8:
        for l in L:
            for n in N2:
                if n==1:
                    obj = obj + (x[c,a,l,n] * GetRev(c, l, 1, 0))
                elif n==2:
                    obj = obj + (x[c,a,l,n] * GetRev(c, l, 1, 1))
                elif n==3:
                    obj = obj + (x[c,a,l,n] * GetRev(c, l, 1, 2))
                elif n==4:
                    obj = obj + (x[c,a,l,n] * GetRev(c, l, 1, 3))
                elif n==5:
                    obj = obj + (x[c,a,l,n] * GetRev(c, l, 2, 0))
                elif n==6:
                    obj = obj + (x[c,a,l,n] * GetRev(c, l, 2, 1))
                elif n==7:
                    obj = obj + (x[c,a,l,n] * GetRev(c, l, 2, 2))
                elif n==8:
                    obj = obj + (x[c,a,l,n] * GetRev(c, l, 3, 0))
                
                    
# #Solving the model

model.setParam('MIPGap', 0)
model.setObjective(obj,GRB.MAXIMIZE)
model.update()
model.optimize()

print(SpecificDemand(1,1,1,2,0))

# #Displaying the solution (objective value)
print("Object:" , model.objVal)


#Displaying the optimal solution
for v in model.getVars():
    if v.x != 0:
        print(v.varName, v.x)


#Printing variable name followed by optimal value
end_time = time.time()
duration = end_time - start_time

minutes, seconds = divmod(duration, 60)
print(f"\nThe Gurobi optimization took {int(minutes)} minutes and {seconds:.2f} seconds to complete.")

